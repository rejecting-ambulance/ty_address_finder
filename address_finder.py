import re
import os
import json
import unicodedata
import time
import subprocess
import logging

from openpyxl import load_workbook
#import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service


def setup_chrome_driver():
    options = Options()

    # --- Headless 模式設定 ---
    options.add_argument('--headless=new')  # 啟用新版無頭模式，模擬真實瀏覽器但不開視窗
    options.add_argument('--disable-gpu')   # 關閉 GPU 加速，避免在部分環境下造成錯誤
    options.add_argument('--no-sandbox')    # 解除沙盒限制（Linux/Docker 無權限環境必加）
    options.add_argument('--disable-dev-shm-usage')  # 避免 /dev/shm 空間不足導致崩潰（Docker 常見）
    options.add_argument('--window-size=1920,1080')  # 指定視窗大小，確保頁面元素完整載入可見

    # --- 日誌與自動化提示設定 ---
    # 排除特定開關，以隱藏「Chrome 正在受自動化控制」提示及多餘的 console log
    options.add_experimental_option(
        'excludeSwitches', 
        ['enable-logging', 'enable-automation']
    )

    # 關閉 Chrome 自動化擴展功能（減少被網站偵測的機率）
    options.add_experimental_option('useAutomationExtension', False)

    # --- 系統級日誌抑制設定 ---
    # 將 Chrome 的內部 log 輸出導向無效位置
    os.environ['CHROME_LOG_FILE'] = os.devnull  # Windows 使用 'NUL'
    # os.environ['CHROME_LOG_FILE'] = '/dev/null'  # Linux/Mac 使用 /dev/null

    # 降低 Selenium 與 urllib3 的日誌輸出層級，只顯示警告以上訊息
    logging.getLogger('selenium').setLevel(logging.WARNING)
    logging.getLogger('urllib3').setLevel(logging.WARNING)

    # 建立並回傳 WebDriver 物件
    # Create a Service that sends chromedriver logs to devnull and hides the console window on Windows
    try:
        creationflags = subprocess.CREATE_NO_WINDOW
    except Exception:
        creationflags = 0

    service = Service(log_path=os.devnull)
    # If supported, set creationflags to avoid spawning visible console windows for the driver
    try:
        service.creationflags = creationflags
    except Exception:
        pass

    # Additional runtime flags to reduce Chrome GPU/log noise
    options.add_argument('--log-level=3')
    options.add_argument('--disable-software-rasterizer')

    return webdriver.Chrome(service=service, options=options)

def load_exception_rules(json_path='exception_rules.json'): # 排除特定里鄰規則
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"require_ling": []}

''' # 原本的等待 class 變化函式，改用 wait_mask_cycle
def wait_class_change(driver, element_id, origin_class, old_class, timeout=10):
    
    WebDriverWait(driver, timeout).until(
        lambda d: d.find_element(By.ID, element_id).get_attribute('class') != origin_class
    )
    WebDriverWait(driver, timeout).until(
        lambda d: d.find_element(By.ID, element_id).get_attribute('class') != old_class
    )
'''


def wait_mask_cycle(driver, mask_class='ext-el-mask', timeout=20):
    """
    等待遮罩出現再消失，用於等待查詢完成
    """
    try:
        # Step 1. 等待遮罩出現
        WebDriverWait(driver, timeout/2).until(
            EC.presence_of_element_located((By.CLASS_NAME, mask_class))
        )
        #print("[INFO] 遮罩已出現，開始等待消失...")

    except Exception:   # 查詢遮罩未出現（可能瞬間出現又消失）
        print("[WARN] 查詢遮罩未出現")

    # Step 2. 等待遮罩消失
    WebDriverWait(driver, timeout).until_not(
        EC.presence_of_element_located((By.CLASS_NAME, mask_class))
    )
    #print("[INFO] 遮罩已消失，查詢完成。")


def search_address(driver, wait, address):
    driver.get('https://addressrs.moi.gov.tw/address/index.cfm?city_id=68000')
    address_box = wait.until(EC.presence_of_element_located((By.ID, 'FreeText_ADDR')))
    #submit_button = driver.find_element(By.ID, 'ext-comp-1010')
    submit_button = driver.find_element(By.ID, 'ext-gen51')

    address_box.clear()
    address_box.send_keys(address)
    submit_button.click()
    
    # 原表單wait
    #wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="ext-gen107"]/div[1]/table/tbody/tr/td[2]/div')))
    
    #wait_class_change(driver, 'ext-gen97', 'x-panel-bwrap', 'x-panel-bwrap x-masked-relative x-masked')
    wait_mask_cycle(driver)

    try:
        #result = driver.find_element(By.XPATH, '//*[@id="ext-gen107"]/div/table/tbody/tr/td[2]/div')
        result = driver.find_element(By.XPATH, '//*[@id="ext-gen111"]/div/table/tbody/tr/td[2]/div')
        return result.text.strip()
    except Exception as e:
        print(f"Error finding result: {e}")
        return "找不到結果"

def simplify_address(address):  # 查詢前地址簡化
    """
    簡化輸入地址，供後續查詢用。

    處理流程重點：
    - 進函式時會先將全形數字轉成半形（例如 '三〇一' → '三0一' 再轉為 '301'）。
    - 移除「里」與「鄰」段（依規則），並擷取「號」之後的後綴作為 suffix。
    - 將地址內的 '-' 轉為 '之'。
    - 在「路/街 + 阿拉伯數字 + 段」情況下，會把 1~9 的阿拉伯數字轉成中文段號（例如 '1段' → '一段'）。
    - 在「路/街 ... 號」情況下，會把緊接於 `號` 前的中文數字逐字轉為阿拉伯數字（包括 '零'、'〇'、'一'~'九'，例如 '三〇一號' → '301號'；但不做單位換算，像 '三百零一號' 會保留原樣）。
    - 若路/街與號之間存在其他文字（例如 '段'），仍會嘗試將緊接在『號』前的中文數字逐字轉為阿拉伯數字。
    - 出函式前會再次將簡化結果中的全形數字轉為半形，回傳格式為 `(original_address, simplified_address, suffix)`。

    備註：本函式僅做逐字對應的中文→阿拉伯數字轉換（非數值運算），如需把含單位的中文數字（十、百、千等）解析成整數，請告知以啟用單位解析。
    """
    original_address = address  # 保留原始輸入

    # 進函式時先將全形數字轉為半形，方便後續處理
    address = fullwidth_to_halfwidth(address)

    # 移除「里」與「鄰」段
    address = re.sub(r'([\u4e00-\u9fff]{1,5}區)[\u4e00-\u9fff]{1,2}里', r'\1', address)
    address = re.sub(r'\d{1,3}鄰', '', address)

    # 處理號後的尾端文字
    split_chars = ['號', '及', '、', '.']
    split_indices = [(address.find(c), c) for c in split_chars if address.find(c) != -1]

    if split_indices:
        split_indices.sort()
        index, char = split_indices[0]

        if char == '號':
            simplified = address[:index + 1]
            suffix = address[index + 1:]
        else:
            simplified = address[:index]
            suffix = address[index:]
    else:
        simplified = address
        suffix = ''

    # 1) 街/路 + 阿拉伯數字(1~2位) + 段 -> 將阿拉伯數字轉為中文段號（支援到十位）
    arabic_digits_map = {0: '零', 1: '一', 2: '二', 3: '三', 4: '四', 5: '五', 6: '六', 7: '七', 8: '八', 9: '九'}

    def arabic_to_chinese_section(n: int) -> str:
        # 支援 1..99 的轉換（十位處理）
        if n <= 0:
            return ''
        if n < 10:
            return arabic_digits_map[n]
        tens, ones = divmod(n, 10)
        if tens == 1:
            # 10..19 -> 十, 十一, 十二...
            return '十' + (arabic_digits_map[ones] if ones else '')
        else:
            return arabic_digits_map[tens] + '十' + (arabic_digits_map[ones] if ones else '')

    def _road_digit_to_chinese(m):
        road = m.group(1)
        num_s = m.group(2)
        # 移除前導零
        num_s = num_s.lstrip('0')
        if not num_s:
            return f"{road}0段"
        n = int(num_s)
        if n >= 1 and n <= 99:
            return f"{road}{arabic_to_chinese_section(n)}段"
        else:
            return f"{road}{num_s}段"

    simplified = re.sub(r'([\u4e00-\u9fff]+(?:路|街))0*([1-9]\d?)段', _road_digit_to_chinese, simplified)

    # 2) 街/路 + 中文數字 + 號 -> 將中文數字逐字對應為阿拉伯數字（不做單位換算）
    char_to_digit = {'零': '0', '〇': '0', '一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
                     '六': '6', '七': '7', '八': '8', '九': '9'}

    def chinese_to_arabic(s: str) -> str:
        # 若包含單位則解析單位（支援到千位），否則作逐字映射
        unit_chars = set('十百千')
        if any(ch in unit_chars for ch in s):
            digits_map = {'零': 0, '〇': 0, '一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
                          '六': 6, '七': 7, '八': 8, '九': 9}
            unit_map = {'千': 1000, '百': 100, '十': 10}
            total = 0
            num = 0
            for ch in s:
                if ch in digits_map:
                    num = digits_map[ch]
                elif ch in unit_map:
                    unit_val = unit_map[ch]
                    if num == 0:
                        num = 1
                    total += num * unit_val
                    num = 0
                else:
                    # 非中文數字或單位，跳過
                    num = 0
            total += num
            return str(total)
        else:
            return ''.join(char_to_digit.get(ch, ch) for ch in s)

    def _road_chinese_to_digit(m):
        road = m.group(1)
        chs = m.group(2)
        arabic = chinese_to_arabic(chs)
        return f"{road}{arabic}號"

    simplified = re.sub(r'([\u4e00-\u9fff]+(?:路|街))([零〇一二三四五六七八九十百千]+)號', _road_chinese_to_digit, simplified)

    # 若路/街 和 號 之間有其他文字（例如「段」），也嘗試把緊接在「號」前的中文數字逐字轉為阿拉伯數字
    def _convert_between_road_and_hao(m):
        prefix = m.group(1)  # 包含路/街及中間文字
        chinese_digits = m.group(2)
        return prefix + chinese_to_arabic(chinese_digits) + '號'

    simplified = re.sub(r'([\u4e00-\u9fff]+(?:路|街).*?)([零〇一二三四五六七八九十百千]+)號', _convert_between_road_and_hao, simplified)
    # 出函式前再確保數字為半形並回傳
    simplified = fullwidth_to_halfwidth(simplified)
    suffix = fullwidth_to_halfwidth(suffix)

    # 號前的數字去除前導零，001 -> 1, 016 -> 16, 010 -> 10
    simplified = re.sub(r"(\d+)號", lambda m: str(int(m.group(1))) + '號', simplified)
    suffix = re.sub(r"(\d+)號", lambda m: str(int(m.group(1))) + '號', suffix)
    #print(f"原地址: {original_address}, 簡化地址: {simplified}, 後綴: {suffix}")
    return original_address.strip(), simplified.strip(), suffix.strip()


def fullwidth_to_halfwidth(text):
    '''
        全形轉半形
    '''
    half_text = ''
    for char in text:
        code = ord(char)
        if code == 0x3000:
            code = 0x0020
        elif 0xFF01 <= code <= 0xFF5E:
            code -= 0xFEE0
        half_text += chr(code)
    return half_text

def format_simplified_address(addr):
    '''
    結果格式化：
    1. 數字轉半形
    2. 去除空格
    3. 將「-」轉回「之」、「,」轉回「，」
    4. 去除「0」開頭的鄰編號，如 003鄰 ➜ 3鄰
    5. 阿拉伯數字轉中文段號（1~9段）
    
    '''    
    
    # 數字轉半形
    addr = fullwidth_to_halfwidth(addr)
    addr = addr.replace(' ', '')  # 去除空格
    addr = addr.replace('-', '之')  # 將「-」轉回「之」
    addr = addr.replace(',', '，')  # 半形「,」轉回「，」

    # 去除「0」開頭的鄰編號，如 003鄰 ➜ 3鄰
    addr = re.sub(r'(\D)0*(\d+)鄰', r'\1\2鄰', addr)

    # 號前的數字去除前導零，001 -> 1, 016 -> 16, 010 -> 10
    addr = re.sub(r'(\d+)號', lambda m: str(int(m.group(1))) + '號', addr)

    # 阿拉伯數字轉中文段號（1~9段）
    num_to_chinese = {'1': '一', '2': '二', '3': '三', '4': '四', '5': '五',
                      '6': '六', '7': '七', '8': '八', '9': '九'}

    def replace_road_section(match):
        num = match.group(1)
        return num_to_chinese.get(num, num) + '段'

    addr = re.sub(r'(\d)段', replace_road_section, addr)

    return addr.strip()

EXCEPTION_RULES = load_exception_rules()
#print("Loaded Exception Rules:", EXCEPTION_RULES)

def remove_ling_with_condition(full_address):
    # 名單內的里，保留鄰

    for special_li in EXCEPTION_RULES.get("require_ling", []):
        if special_li in full_address:
            return full_address
        
    # 否則執行標準簡化：刪除「里」與「鄰」間文字（含鄰）
    return re.sub(r'(里).*?鄰', r'\1', full_address)


def process_no_result_address(original_address):
    """
    處理查無結果的地址：如果原地址有「里」，直接放到「不含鄰的地址」欄
    """
    if "里" in original_address:
        if "桃園市" not in original_address:
            original_address = f'桃園市{original_address}'
        # 非例外里，只要有「里」就保留
        return original_address
    else:
        return "查詢失敗"

def visual_len(text):
    """ 計算文字的實際顯示寬度 """
    width = 0
    for ch in text:
        if unicodedata.east_asian_width(ch) in ('F', 'W'):
            width += 2  # 全形、寬字元
        else:
            width += 1  # 半形
    return width

def pad_text(text, target_width):
    """ 補足空格讓文字達到指定寬度 """
    pad_len = target_width - visual_len(text)
    return text + ' ' * max(pad_len, 0)

def read_addresses(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    return [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]

def jurisdiction_check(data_path = "address_data.xlsx",jurisdiction_path='責任區.xlsx'):
    # 取得程式執行目錄
    folder_path = os.getcwd()

    # 檔案名稱
    target_file = os.path.join(folder_path, jurisdiction_path)
    source_file = os.path.join(folder_path, data_path)

    if os.path.exists(target_file) and os.path.exists(source_file):
        # 開啟 address_data.xlsx
        wb_source = load_workbook(source_file)
        ws_source = wb_source.active  # 假設資料在第一個工作表

        # 開啟責任區.xlsx
        wb_target = load_workbook(target_file)
        
        # 如果工作表不存在就建立
        if "程式用" in wb_target.sheetnames:
            ws_target = wb_target["程式用"]
        else:
            ws_target = wb_target.create_sheet("程式用")
        
        # 清空目標工作表的 A~D 欄舊資料（可選）
        for row in ws_target.iter_rows(min_col=1, max_col=4, max_row=ws_target.max_row):
            for cell in row:
                cell.value = None

        # 取得 source A~D 欄資料並寫入
        for row_idx, row in enumerate(ws_source.iter_rows(min_col=1, max_col=4), start=1):
            for col_idx, cell in enumerate(row, start=1):
                ws_target.cell(row=row_idx, column=col_idx, value=cell.value)
        
        # 儲存修改
        wb_target.save(target_file)
        return True
    else:
        None
        #print("資料夾內缺少必要檔案：責任區.xlsx 或 address_data.xlsx")


def main(file_path = "address_data.xlsx", jurisdiction_path = "責任區.xlsx"):

    #df = pd.read_excel(file_path)
    #addresses = df['查詢地址'].tolist()
    addresses = read_addresses(file_path)

    driver = setup_chrome_driver()
    wait = WebDriverWait(driver, 10)

    # 開啟 Excel 用來逐筆寫入
    wb = load_workbook(file_path)
    ws = wb.active

    max_len = 50  # 用來對齊箭頭

    for i, address in enumerate(addresses, start=1):
        if not address or str(address).strip() == 'nan':
            print(f"{i:04d}. 空白資料")
            full_address = ""
            simplified = ""
        else:
            time.sleep(1.2)  # 避免查詢過快被擋
            try:
                data_address, shorter_address, last_address = simplify_address(address)
                result_address = search_address(driver, wait, shorter_address)

                if result_address == "找不到結果":
                    
                    simplified = process_no_result_address(data_address)
                    simplified = remove_ling_with_condition(simplified)

                    if "里" in simplified:
                        full_address = "查無結果，使用原里鄰"
                        output = f"{i:04d}. {pad_text(address, max_len)} → {simplified}(查無結果，使用原里鄰)"
                    else:
                        full_address = "查無結果"
                        output = f"{i:04d}. {pad_text(address, max_len)} → 查無結果"

                    print(output)
                else:
                    full_address = f'桃園市{result_address}{last_address}'
                    full_address = fullwidth_to_halfwidth(full_address)
                    full_address = full_address.replace(',', '，')

                    simplified = remove_ling_with_condition(full_address)

                    output = f"{i:04d}. {pad_text(address, max_len)} → {full_address}"
                    print(output)

            except Exception as e:
                print(f"{i:04d}. {pad_text(address, max_len)} → 查詢失敗")
                full_address = "查詢失敗"
                simplified = process_no_result_address(address)

        # 統一簡化地址格式
        formatted_simplified = format_simplified_address(simplified)

        # 寫入第 i+1 列（因為 Excel 有標題列）
        ws.cell(row=i+1, column=1, value=i)  # A欄流水號
        ws.cell(row=i+1, column=3, value=full_address)  # C欄：完整地址
        ws.cell(row=i+1, column=4, value=formatted_simplified)  # D欄：不含鄰的地址

        # 每筆處理完就儲存一次
        wb.save(file_path)

    driver.quit()

    if(jurisdiction_check()):
        print(f"✅ 責任區已更新，請查看：{jurisdiction_path}")
        os.startfile(jurisdiction_path)
    else:
        print(f"✅ 查詢結束，請查看：{file_path}")
        os.startfile(file_path)


if __name__ == '__main__':
    file_path = 'address_data.xlsx'
    jurisdiction_path = '責任區.xlsx'
    main(file_path, jurisdiction_path)
    
    